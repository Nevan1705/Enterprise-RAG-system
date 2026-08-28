"""
================================================================================
Project     : Enterprise RAG System v3
Module      : backend/app/services/excel_export.py
Author      : Enterprise AI Engineering Team
Date        : 2026-08-28
Description : Excel Export Engine for QA Test Cases. Converts validated testcase
              dictionaries into professionally styled, wrapped, and formatted
              Microsoft Excel (.xlsx) workbooks using pandas and openpyxl.
================================================================================
"""
import os
import uuid
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Ensure persistent downloads directory exists
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "..", "downloads")
os.makedirs(DOWNLOADS_DIR, exist_ok=True)


# -----------------------------------------------------------------------------
# Excel Test Case Workbook Generator
# -----------------------------------------------------------------------------
def export_testcases_to_excel(testcases: list[dict]) -> str:
    """
    Exports a list of test case dictionaries to a stylized Excel (.xlsx) file.
    
    Args:
        testcases (list[dict]): List of validated TestCaseItem dictionaries.
        
    Returns:
        str: Filename of the generated spreadsheet in the downloads directory.
    """
    # 1. Format test steps as numbered multi-line text
    formatted_data = []
    for tc in testcases:
        item = tc.copy()
        if isinstance(item.get("test_steps"), list):
            item["test_steps"] = "\n".join(f"{i+1}. {step}" for i, step in enumerate(item["test_steps"]))
        formatted_data.append(item)
        
    df = pd.DataFrame(formatted_data)
    
    # 2. Rename columns to human-readable enterprise titles
    column_mapping = {
        "test_case_id": "Test Case ID",
        "requirement_id": "Requirement ID",
        "test_scenario": "Test Scenario",
        "test_case_description": "Test Case Description",
        "preconditions": "Preconditions",
        "test_steps": "Test Steps",
        "test_data": "Test Data",
        "expected_result": "Expected Result"
    }
    df = df.rename(columns=column_mapping)
    
    # 3. Generate unique filename
    filename = f"testcases_{uuid.uuid4().hex[:8]}.xlsx"
    filepath = os.path.join(DOWNLOADS_DIR, filename)
    
    # 4. Write Excel workbook with custom openpyxl styling
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Test Cases')
        
        workbook = writer.book
        worksheet = writer.sheets['Test Cases']
        
        # Header Styling: Bold text and clean font
        header_font = Font(name='Calibri', size=11, bold=True, color='000000')
        header_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
        # Column Dimensioning & Cell Wrapping
        for col_idx, column in enumerate(worksheet.columns, 1):
            col_letter = get_column_letter(col_idx)
            worksheet.column_dimensions[col_letter].width = 28
            
            for cell in column:
                if cell.row > 1:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
    return filename
